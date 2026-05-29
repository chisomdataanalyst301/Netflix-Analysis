"""
Netflix Content Library — Data Cleaning & Transformation Script
================================================================
Author  : Data Analyst
Project : Netflix Content Library Analysis
Date    : May 2026

Description:
    Reads the raw Netflix_Data.xlsx file, performs full data cleaning,
    adds 8 calculated columns, and exports a Power BI-ready Excel file
    with correct column types and formatting.

Usage:
    python data_cleaning.py

Input  : data/Netflix_Data.xlsx
Output : data/Netflix_PowerBI_Fixed.xlsx
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime
import os

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
INPUT_FILE  = "data/Netflix_Data.xlsx"
OUTPUT_FILE = "data/Netflix_PowerBI_Fixed.xlsx"


def load_data(path: str) -> pd.DataFrame:
    """Load raw Excel file and strip column name whitespace."""
    print(f"📂 Loading: {path}")
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()
    print(f"   Loaded {len(df):,} rows × {len(df.columns)} columns")
    return df


def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Full cleaning pipeline."""
    print("\n🧹 Cleaning data...")

    # ── Remove duplicates ──────────────────────────────────────────────
    before = len(df)
    df = df.drop_duplicates(subset=["Title", "Type"], keep="first")
    dupes = before - len(df)
    print(f"   Removed {dupes} duplicates")

    # ── Fill missing values ────────────────────────────────────────────
    df["Director"]     = df["Director"].fillna("Unknown")       if "Director"     in df.columns else df
    df["Country"]      = df["Country"].fillna("Not Specified")
    df["Rating"]       = df["Rating"].fillna("NR")
    df["Release Year"] = df["Release Year"].fillna(df["Release Year"].median())
    df["Release Year"] = df["Release Year"].astype(int)

    print(f"   Nulls remaining: {df.isnull().sum().sum()}")

    # ── Force Date Added to plain string YYYY-MM-DD ────────────────────
    df["Date Added"] = df["Date Added"].astype(str).str.strip()

    return df.reset_index(drop=True)


def add_calculated_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add all 8 derived columns."""
    print("\n📐 Adding calculated columns...")

    # Date-based columns
    df["Year Added"]    = df["Date Added"].str[:4].astype(int)
    df["Month Num"]     = df["Date Added"].str[5:7].astype(int)
    df["Month Name"]    = pd.to_datetime(df["Date Added"]).dt.strftime("%B")
    df["Quarter Added"] = "Q" + pd.to_datetime(df["Date Added"]).dt.quarter.astype(str)

    # Release era (5-year buckets)
    df["Release Era"] = pd.cut(
        df["Release Year"],
        bins=[1999, 2004, 2009, 2014, 2019, 2024, 2100],
        labels=["2000–2004", "2005–2009", "2010–2014",
                "2015–2019", "2020–2024", "2025+"]
    ).astype(str)

    # Movie duration bucket
    def duration_bucket(row):
        if row["Type"] == "TV Show":
            return "TV Show"
        m = row["Duration Minutes"]
        if pd.isna(m):
            return "Unknown"
        if m < 90:   return "Under 90 min"
        if m < 110:  return "90–110 min"
        if m < 130:  return "110–130 min"
        if m < 150:  return "130–150 min"
        if m < 170:  return "150–170 min"
        return "170+ min"

    df["Duration Bucket"] = df.apply(duration_bucket, axis=1)

    # Rating group
    adult  = ["TV-MA", "R", "NC-17"]
    teen   = ["TV-14", "PG-13"]
    family = ["PG", "G", "TV-PG", "TV-G", "TV-Y", "TV-Y7"]

    def rating_group(r):
        if r in adult:  return "Adult"
        if r in teen:   return "Teen"
        if r in family: return "Family"
        return "Unrated"

    df["Rating Group"] = df["Rating"].apply(rating_group)

    # Churn label
    df["Churn Label"] = df["Type"]

    print(f"   Added 8 columns → total: {len(df.columns)} columns")
    return df


def export_to_excel(df: pd.DataFrame, output_path: str):
    """Export clean dataframe to formatted Excel file."""
    print(f"\n💾 Exporting to: {output_path}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Netflix Data"
    ws.sheet_view.showGridLines = False

    # Styles
    RED  = PatternFill("solid", start_color="E50914", end_color="E50914")
    ALT  = PatternFill("solid", start_color="F9F9F9", end_color="F9F9F9")
    thin = Side(style="thin", color="E0E0E0")
    BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
    WBOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    DREG  = Font(name="Calibri", color="1A1A1A", size=9)
    CTR   = Alignment(horizontal="center", vertical="center")
    LEFT  = Alignment(horizontal="left",   vertical="center")

    headers = list(df.columns)
    date_col_idx = headers.index("Date Added") + 1

    # Write headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill      = RED
        cell.font      = WBOLD
        cell.alignment = CTR
        cell.border    = BDR

    # Write data rows
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = ALT if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            # Sanitise value
            if val is None or (isinstance(val, float) and np.isnan(val)):
                val = None
            elif hasattr(val, "item"):
                val = val.item()

            cell = ws.cell(ri, ci, val)
            cell.font      = DREG
            cell.alignment = LEFT
            cell.border    = BDR
            if fill:
                cell.fill = fill

            # Force Date Added to text — prevents Power BI mm-dd-yy corruption
            if ci == date_col_idx:
                cell.number_format = "@"

    # Column widths
    col_widths = {
        "Id": 6, "Title": 22, "Type": 10, "Genre": 14,
        "Release Year": 13, "Rating": 8, "Duration Minutes": 16,
        "Seasons": 9, "Country": 16, "Language": 12,
        "Date Added": 13, "Views": 10, "Estimated Revenue Usd": 22,
        "Avg Watch Time Pct": 18, "Completion Rate Pct": 18,
        "Engagement Score": 16, "Popularity Index": 15,
        "Year Added": 11, "Month Num": 10, "Month Name": 13,
        "Quarter Added": 13, "Release Era": 13, "Duration Bucket": 16,
        "Rating Group": 13, "Churn Label": 12,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 20

    wb.save(output_path)
    print(f"   ✅ Saved {len(df):,} rows × {len(df.columns)} columns")


def print_summary(df: pd.DataFrame):
    """Print a clean summary of the cleaned dataset."""
    print("\n" + "=" * 50)
    print("📊 DATASET SUMMARY")
    print("=" * 50)
    print(f"  Total titles      : {len(df):,}")
    print(f"  Movies            : {(df['Type']=='Movie').sum():,}")
    print(f"  TV Shows          : {(df['Type']=='TV Show').sum():,}")
    print(f"  Countries         : {df['Country'].nunique()}")
    print(f"  Languages         : {df['Language'].nunique()}")
    print(f"  Genres            : {df['Genre'].nunique()}")
    print(f"  Total revenue     : ${df['Estimated Revenue Usd'].sum():,.0f}")
    print(f"  Avg revenue/title : ${df['Estimated Revenue Usd'].mean():,.0f}")
    print(f"  Avg completion    : {df['Completion Rate Pct'].mean():.1f}%")
    print(f"  Avg engagement    : {df['Engagement Score'].mean():.2f}")
    print(f"  Year range        : {df['Year Added'].min()} – {df['Year Added'].max()}")
    print("=" * 50)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("🎬 Netflix Data Cleaning Pipeline")
    print(f"   Started: {datetime.now().strftime('%d %b %Y %H:%M:%S')}\n")

    df = load_data(INPUT_FILE)
    df = clean_data(df)
    df = add_calculated_columns(df)
    print_summary(df)
    export_to_excel(df, OUTPUT_FILE)

    print(f"\n✅ Done! Output saved to: {OUTPUT_FILE}")
    print(f"   Finished: {datetime.now().strftime('%d %b %Y %H:%M:%S')}")
